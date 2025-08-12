import io
import csv
import asyncio
from openpyxl import load_workbook
import xlrd   


async def extract_tabular_text(file_bytes: bytes, file_name: str) -> str:
   
    ext = file_name.lower().split('.')[-1]

    if ext == "xlsx":
        return await asyncio.to_thread(_extract_xlsx, file_bytes)

    elif ext == "xls":
        return await asyncio.to_thread(_extract_xls, file_bytes)

    elif ext == "csv":
        return await asyncio.to_thread(_extract_csv, file_bytes)

    else:
        raise ValueError(f"Unsupported file extension: {ext}")


def _extract_xlsx(file_bytes: bytes) -> str:
    stream = io.BytesIO(file_bytes)
    wb = load_workbook(stream, data_only=True)
    text_parts = []
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        has_data = any(
            cell is not None and str(cell).strip()
            for row in ws.iter_rows(values_only=True)
            for cell in row
        )
        if not has_data:
            continue
        text_parts.append(f"=== {sheet} ===")
        for row in ws.iter_rows(values_only=True):
            row_values = [str(cell) if cell is not None else "" for cell in row]
            text_parts.append("\t".join(row_values))
    return "\n".join(text_parts)


def _extract_xls(file_bytes: bytes) -> str:
    stream = io.BytesIO(file_bytes)
    book = xlrd.open_workbook(file_contents=stream.read())
    text_parts = []
    for sheet in book.sheets():
        has_data = any(
            str(sheet.cell_value(rx, cx)).strip()
            for rx in range(sheet.nrows)
            for cx in range(sheet.ncols)
        )
        if not has_data:
            continue
        text_parts.append(f"=== {sheet.name} ===")
        for rx in range(sheet.nrows):
            row_values = [str(sheet.cell_value(rx, cx)) for cx in range(sheet.ncols)]
            text_parts.append("\t".join(row_values))
    return "\n".join(text_parts)


def _extract_csv(file_bytes: bytes) -> str:
    stream = io.StringIO(file_bytes.decode("utf-8", errors="ignore"))
    reader = list(csv.reader(stream))
    if not any(any(cell.strip() for cell in row) for row in reader):
        return ""
    return "\n".join("\t".join(row) for row in reader)
